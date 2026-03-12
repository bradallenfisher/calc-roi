#!/usr/bin/env python3
"""
Extract formulas and values from ASAP ECC ROI Calculator Excel file.
Run: ./venv/bin/python extract_excel_formulas.py
Output is written to excel_formulas.txt for easy review.
"""

import openpyxl
from pathlib import Path

# Set which file to extract: "original" or "cost_categories"
EXTRACT_FILE = "cost_categories"

XLSX_PATHS = {
    "original": Path(__file__).parent / "ASAP ECC ROI Calculator 20260130 JS.xlsx",
    "cost_categories": Path(__file__).parent / "ROI Cost Categories 20260227 JS.xlsx",
}
XLSX_PATH = XLSX_PATHS.get(EXTRACT_FILE, XLSX_PATHS["original"])
if not XLSX_PATH.exists():
    XLSX_PATH = XLSX_PATHS["original"]
OUTPUT_PATH = Path(__file__).parent / "excel_formulas.txt"


def cell_ref(cell):
    return f"{cell.column_letter}{cell.row}"


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    lines = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        lines.append(f"\n{'='*60}")
        lines.append(f"SHEET: {sheet_name}")
        lines.append("=" * 60)

        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                if cell.value is None:
                    ref = cell_ref(cell)
                    row_data.append(f"{ref}: (empty)")
                elif isinstance(cell.value, str) and cell.value.strip().startswith("="):
                    ref = cell_ref(cell)
                    row_data.append(f"{ref}: FORMULA: {cell.value}")
                else:
                    ref = cell_ref(cell)
                    row_data.append(f"{ref}: {cell.value!r}")

            row_str = " | ".join(row_data)
            if row_str.strip() and "empty" not in row_str.lower() or "FORMULA" in row_str:
                lines.append(row_str)

    # Also load with data_only=True to show computed values for formula cells
    wb_values = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    lines.append("\n\n" + "=" * 60)
    lines.append("COMPUTED VALUES (for formula cells)")
    lines.append("=" * 60)

    for sheet_name in wb_values.sheetnames:
        sheet = wb_values[sheet_name]
        lines.append(f"\n--- {sheet_name} ---")
        for row in sheet.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value is not None:
                    ref = cell_ref(cell)
                    lines.append(f"  {ref}: {cell.value}")

    output = "\n".join(lines)
    OUTPUT_PATH.write_text(output, encoding="utf-8")
    print(f"Written to {OUTPUT_PATH}")
    print(output[:4000])
    if len(output) > 4000:
        print(f"\n... (truncated, full output in {OUTPUT_PATH})")


if __name__ == "__main__":
    main()
